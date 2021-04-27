import time
from time import strftime

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side

wb = Workbook()

ws = wb.active


def criarPlanilha(individuo):
    constLinha = 4
    constColuna = 3
    for periodo in range(6):
        lista = []
        for linha in range(4):
            for coluna in range(5):
                celulaLinha = linha + constLinha + periodo * 7
                celunaColuna = coluna + constColuna
                aula = aulaInfo(individuo[periodo][coluna + linha * 5])
                celula = ws.cell(row=celulaLinha, column=celunaColuna, value=aula)
                celula.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')

                ws.row_dimensions[celula.row].height = 60
                ws.column_dimensions[str(celula.column_letter)].width = 38

                if individuo[periodo][coluna + linha * 5]['disciplina'] is not None:
                    disciplina = individuo[periodo][coluna + linha * 5]['disciplina']['nome']
                    turma = individuo[periodo][coluna + linha * 5]['turma']

                    cor, lista = cellCor(disciplina, turma, lista)
                else:
                    cor = 'D3D3D3'

                ws[celula.coordinate].fill = PatternFill(start_color=cor, fill_type='solid')
                ws[celula.coordinate].border = Border(left=Side(border_style='medium',
                                                                color='000000'),
                                                      right=Side(border_style='medium',
                                                                 color='000000'),
                                                      top=Side(border_style='medium',
                                                               color='000000'),
                                                      bottom=Side(border_style='medium',
                                                                  color='000000'))

    for periodo in range(6):
        for linha in range(4):
            celulaLinha = linha + constLinha + periodo * 7
            celula = ws.cell(row=celulaLinha, column=2, value=horarios(linha))
            celula.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')

            # ws.row_dimensions[celula.row].height = 60
            ws.column_dimensions[str(celula.column_letter)].width = 15

            ws[celula.coordinate].fill = PatternFill(start_color='808080', fill_type='solid')
            ws[celula.coordinate].border = Border(left=Side(border_style='medium',
                                                            color='000000'),
                                                  right=Side(border_style='medium',
                                                             color='000000'),
                                                  top=Side(border_style='medium',
                                                           color='000000'),
                                                  bottom=Side(border_style='medium',
                                                              color='000000'))

            # cor = '%02X%02X%02X' % (random.randint(0, 255), random.randint(0, 255), random.randint(0, 255))
            # ws[celula.coordinate].fill = PatternFill(start_color=cor, fill_type='solid')

    for periodo in range(6):
        for coluna in range(5):
            celulaColuna = coluna + constColuna
            celulaLinha = constLinha + periodo * 7 - 1

            celula = ws.cell(row=celulaLinha, column=celulaColuna, value=dias(coluna))

            ws.row_dimensions[celula.row].height = 30

            celula.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')

            ws[celula.coordinate].fill = PatternFill(start_color='808080', fill_type='solid')
            ws[celula.coordinate].border = Border(left=Side(border_style='medium',
                                                            color='000000'),
                                                  right=Side(border_style='medium',
                                                             color='000000'),
                                                  top=Side(border_style='medium',
                                                           color='000000'),
                                                  bottom=Side(border_style='medium',
                                                              color='000000'))

            # ws.row_dimensions[celula.row].height = 60
            # ws.column_dimensions[str(celula.column_letter)].width = 15

            # ws[celula.coordinate].fill = PatternFill(start_color=cor, fill_type='solid')

    for periodo in range(6):
        celulaLinha = constLinha + periodo * 7 - 2
        inicialCelula = ws.cell(row=celulaLinha, column=constColuna)
        finalCelula = ws.cell(row=celulaLinha, column=constColuna + 4)

        mergedCelula = inicialCelula.coordinate + ':' + finalCelula.coordinate

        ws.merge_cells(mergedCelula)

        ws[inicialCelula.coordinate] = str((periodo + 1)) + '° Periodo '
        celula = ws[inicialCelula.coordinate]

        ws.row_dimensions[celula.row].height = 30

        celula.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')

        ws[celula.coordinate].fill = PatternFill(start_color='808080', fill_type='solid')
        ws[celula.coordinate].border = Border(left=Side(border_style='medium',
                                                        color='000000'),
                                              right=Side(border_style='medium',
                                                         color='000000'),
                                              top=Side(border_style='medium',
                                                       color='000000'),
                                              bottom=Side(border_style='medium',
                                                          color='000000'))

    wb.save('./Grade de Horarios/grade_de_horarios ' + data() + '.xlsx')


def aulaInfo(aula):
    if aula['disciplina'] is not None:
        disciplina = aula['disciplina']['nome']
        professor = aula['professor']['nome']
        turma = aula['turma']

        info = 'Disciplina: ' + disciplina + '\nProfessor: ' + professor + '\nTurma: ' + str((1 + turma)) + ''
    else:
        info = 'Sem Aula'
    return info


def horarios(linha):
    if linha == 0:
        return '7:55 - 9:45'
    elif linha == 1:
        return '9:45 - 11-35'
    elif linha == 2:
        return '13:55 - 15:45'
    else:
        return '15:45 - 17:35'


def dias(coluna):
    if coluna == 0:
        return 'Segunda-Feira'
    elif coluna == 1:
        return 'Terça-Feira'
    elif coluna == 2:
        return 'Quarta-Feira'
    elif coluna == 3:
        return 'Quinta-Feira'
    else:
        return 'Sexta-Feira'


def data():
    return strftime('%d-%m-%Y %H-%M-%S', time.localtime())


def cellCor(disciplina, turma, lista):
    cor = ['FF1493', '8A2BE2', 'D2691E', '00FF00', '00FFFF', 'EE82EE', '4169E1', 'B22222', 'FFD700', 'B0E0E6', '3CB371',
           'BDB76B']

    for i in range(len(lista)):
        if disciplina == lista[i][0] and turma == lista[i][1]:
            return [lista[i][2], lista]

    lista.append((disciplina, turma, cor[len(lista)]))
    return [cor[len(lista) - 1], lista]
